from __future__ import annotations

from openpyxl import load_workbook
from app.models import TableMap, FieldMap, ValueMap


def _sheet_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) for h in rows[0]]
    out = []
    for r in rows[1:]:
        out.append({headers[i]: r[i] for i in range(len(headers))})
    return out


def load_mapping(path: str) -> tuple[list[TableMap], list[FieldMap], list[ValueMap]]:
    wb = load_workbook(path)
    table_rows = _sheet_dicts(wb["table_map"])
    field_rows = _sheet_dicts(wb["field_map"])

    value_rows = []
    if "value_map" in wb.sheetnames:
        value_rows = _sheet_dicts(wb["value_map"])

    table_maps = [TableMap(**r) for r in table_rows]
    field_maps = [FieldMap(**r) for r in field_rows]
    value_maps = [ValueMap(**r) for r in value_rows]
    return table_maps, field_maps, value_maps


def field_map_for_table(field_maps: list[FieldMap], left_table: str, right_table: str) -> list[FieldMap]:
    return [
        f for f in field_maps
        if f.left_table == left_table and f.right_table == right_table and f.compare
    ]


def key_map_for_table(field_maps: list[FieldMap], left_table: str, right_table: str) -> list[FieldMap]:
    return [
        f for f in field_maps
        if f.left_table == left_table and f.right_table == right_table and f.is_key
    ]
