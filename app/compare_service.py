from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from app.db import get_table_columns, run_query
from app.mapping_store import field_map_for_table, key_map_for_table
from app.models import FieldMap


def _quote_ident(name: str) -> str:
    # ANSI SQL identifier quoting with escaped double-quotes.
    # Allows spaces, reserved keywords, plus signs, etc. safely.
    return '"' + name.replace('"', '""') + '"'


def _quote_table(table_name: str) -> str:
    # Quote each dotted identifier part: schema.table -> "schema"."table"
    return ".".join(_quote_ident(part) for part in table_name.split("."))


def _looks_datetime_field(field_name: str) -> bool:
    n = field_name.lower()
    tokens = ("date", "time", "timestamp", "datetime", "_dt", "_dttm")
    return any(t in n for t in tokens)


def _normalize_datetime_value(v: Any):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day, 0, 0, 0).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # try common datetime/date layouts
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
            "%m/%d/%Y",
        ):
            try:
                dt = datetime.strptime(s, fmt)
                if fmt in ("%Y-%m-%d", "%m/%d/%Y"):
                    dt = datetime(dt.year, dt.month, dt.day, 0, 0, 0)
                return dt.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                pass
        # ISO fallbacks
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return s
    return v


def _normalize(v: Any, *, field_name: str | None = None, trim_strings: bool = True, nullish_equal: bool = True, number_precision: int = 6):
    if v is None:
        return None
    if field_name and _looks_datetime_field(field_name):
        return _normalize_datetime_value(v)
    if isinstance(v, str):
        vv = v.strip() if trim_strings else v
        if nullish_equal and vv == "":
            return None
        return vv
    if isinstance(v, float):
        return round(v, number_precision)
    return v


def _rows_to_dicts(cols: list[str], rows: list[tuple]) -> list[dict[str, Any]]:
    return [dict(zip(cols, r)) for r in rows]


def _unique_keep_order(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for v in values:
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out


def _sheet_name(name: str) -> str:
    safe = name.replace("/", "_").replace("\\", "_").replace(".", "_")
    return safe[:31] if len(safe) > 31 else safe


def _safe_table_label(table_name: str) -> str:
    # keep only table part (after schema), remove periods/slashes/backslashes
    base = table_name.split(".")[-1]
    return base.replace("/", "_").replace("\\", "_").replace(".", "_")


def _write_table(ws, rows: list[dict[str, Any]], headers: list[str] | None = None):
    if headers is None:
        headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])


def compare_tables(
    left_conn: dict,
    right_conn: dict,
    left_table: str,
    right_table: str,
    all_field_maps: list[FieldMap],
    employee_id: str | None = None,
    sample_employee_ids: bool = False,
    sample_size: int = 100,
    god_mode: bool = False,
    trim_strings: bool = True,
    nullish_equal: bool = True,
    number_precision: int = 6,
    output_dir: str = "output",
) -> dict[str, Any]:
    pair_field_maps = [
        f for f in all_field_maps
        if f.left_table == left_table and f.right_table == right_table
    ]
    compare_maps = field_map_for_table(all_field_maps, left_table, right_table)
    key_maps = key_map_for_table(all_field_maps, left_table, right_table)

    if not compare_maps:
        raise ValueError("No compare field mappings for selected table pair")
    if not key_maps:
        raise ValueError("No key mappings for selected table pair")

    left_schema_cols = get_table_columns(left_conn, left_table)
    right_schema_cols = get_table_columns(right_conn, right_table)

    mapped_left_fields = {f.left_field for f in pair_field_maps}
    mapped_right_fields = {f.right_field for f in pair_field_maps}
    left_only_fields = [c for c in left_schema_cols if c not in mapped_left_fields]
    right_only_fields = [c for c in right_schema_cols if c not in mapped_right_fields]

    left_key_fields = [k.left_field for k in key_maps]
    right_key_fields = [k.right_field for k in key_maps]

    left_compare_fields = [f.left_field for f in compare_maps]
    right_compare_fields = [f.right_field for f in compare_maps]

    left_select_fields = _unique_keep_order(left_key_fields + left_compare_fields)
    right_select_fields = _unique_keep_order(right_key_fields + right_compare_fields)

    left_cols_sql = ", ".join(_quote_ident(c) for c in left_select_fields)
    right_cols_sql = ", ".join(_quote_ident(c) for c in right_select_fields)

    left_sql = f"SELECT {left_cols_sql} FROM {_quote_table(left_table)}"
    right_sql = f"SELECT {right_cols_sql} FROM {_quote_table(right_table)}"

    left_params: tuple | None = None
    right_params: tuple | None = None

    lk_emp = next((k for k in left_key_fields if k.lower() == "employee_id"), None)
    rk_emp = next((k for k in right_key_fields if k.lower() == "employee_id"), None)

    if employee_id:
        if lk_emp and rk_emp:
            left_sql += f" WHERE {_quote_ident(lk_emp)} = %s"
            right_sql += f" WHERE {_quote_ident(rk_emp)} = %s"
            left_params = (employee_id,)
            right_params = (employee_id,)
    elif sample_employee_ids and lk_emp and rk_emp:
        sample_n = max(1, int(sample_size))
        sample_sql = (
            f"SELECT DISTINCT {_quote_ident(lk_emp)} "
            f"FROM {_quote_table(left_table)} "
            f"WHERE {_quote_ident(lk_emp)} IS NOT NULL "
            f"ORDER BY RANDOM() LIMIT {sample_n}"
        )
        s_cols, s_rows, _ = run_query(left_conn, sample_sql)
        sampled_ids = [r[s_cols.index(lk_emp)] for r in s_rows] if s_rows else []

        if sampled_ids:
            placeholders = ", ".join(["%s"] * len(sampled_ids))
            left_sql += f" WHERE {_quote_ident(lk_emp)} IN ({placeholders})"
            right_sql += f" WHERE {_quote_ident(rk_emp)} IN ({placeholders})"
            left_params = tuple(sampled_ids)
            right_params = tuple(sampled_ids)

    l_cols, l_rows, l_sec = run_query(left_conn, left_sql, left_params)
    r_cols, r_rows, r_sec = run_query(right_conn, right_sql, right_params)

    left_data = _rows_to_dicts(l_cols, l_rows)
    right_data = _rows_to_dicts(r_cols, r_rows)

    def lkey(row: dict[str, Any]):
        return tuple(_normalize(row.get(k), field_name=k, trim_strings=trim_strings, nullish_equal=nullish_equal, number_precision=number_precision) for k in left_key_fields)

    def rkey(row: dict[str, Any]):
        return tuple(_normalize(row.get(k), field_name=k, trim_strings=trim_strings, nullish_equal=nullish_equal, number_precision=number_precision) for k in right_key_fields)

    left_ix = {lkey(row): row for row in left_data}
    right_ix = {rkey(row): row for row in right_data}

    left_keys = set(left_ix)
    right_keys = set(right_ix)

    only_left_rows = [left_ix[k] for k in sorted(left_keys - right_keys)]
    only_right_rows = [right_ix[k] for k in sorted(right_keys - left_keys)]

    def _norm_row(row: dict[str, Any], fields: list[str]) -> dict[str, Any]:
        return {
            f: _normalize(row.get(f), field_name=f, trim_strings=trim_strings, nullish_equal=nullish_equal, number_precision=number_precision)
            for f in fields
        }

    only_left_rows_norm = [_norm_row(r, left_select_fields) for r in only_left_rows]
    only_right_rows_norm = [_norm_row(r, right_select_fields) for r in only_right_rows]
    left_data_norm = [_norm_row(r, left_select_fields) for r in left_data]
    right_data_norm = [_norm_row(r, right_select_fields) for r in right_data]

    field_differences_rows: list[dict[str, Any]] = []
    summary_counts: dict[tuple[str, str], int] = {}

    for k in sorted(left_keys & right_keys):
        lrow = left_ix[k]
        rrow = right_ix[k]

        key_payload = {lf: _normalize(lrow.get(lf), field_name=lf, trim_strings=trim_strings, nullish_equal=nullish_equal, number_precision=number_precision) for lf in left_key_fields}

        for fm in compare_maps:
            lv_raw = lrow.get(fm.left_field)
            rv_raw = rrow.get(fm.right_field)
            lv = _normalize(lv_raw, field_name=fm.left_field, trim_strings=trim_strings, nullish_equal=nullish_equal, number_precision=number_precision)
            rv = _normalize(rv_raw, field_name=fm.right_field, trim_strings=trim_strings, nullish_equal=nullish_equal, number_precision=number_precision)
            if lv != rv:
                out = {
                    **key_payload,
                    "field_name": f"{fm.left_field} -> {fm.right_field}",
                    "left_value": lv,
                    "right_value": rv,
                }
                field_differences_rows.append(out)
                key = (fm.left_field, fm.right_field)
                summary_counts[key] = summary_counts.get(key, 0) + 1

    summary_rows = [
        {"left_field": lf, "right_field": rf, "difference_count": cnt}
        for (lf, rf), cnt in sorted(summary_counts.items(), key=lambda x: (-x[1], x[0][0], x[0][1]))
    ]

    # Output workbook
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%H%M")
    right_label = _safe_table_label(right_table)
    out_path = out_dir / f"compare_{right_label}_{ts}.xlsx"

    wb = Workbook()
    ws_left_only = wb.active
    ws_left_only.title = _sheet_name("only in left")
    _write_table(ws_left_only, only_left_rows_norm, headers=left_select_fields)

    ws_right_only = wb.create_sheet(_sheet_name("only in right"))
    _write_table(ws_right_only, only_right_rows_norm, headers=right_select_fields)

    ws_diffs = wb.create_sheet(_sheet_name("field differences"))
    diff_headers = left_key_fields + ["field_name", "left_value", "right_value"]
    _write_table(ws_diffs, field_differences_rows, headers=diff_headers)

    ws_summary = wb.create_sheet(_sheet_name("field difference summary"))
    _write_table(ws_summary, summary_rows, headers=["left_field", "right_field", "difference_count"])

    ws_left_fields = wb.create_sheet(_sheet_name("left only fields"))
    _write_table(ws_left_fields, [{"left_field": f} for f in left_only_fields], headers=["left_field"])

    ws_right_fields = wb.create_sheet(_sheet_name("right only fields"))
    _write_table(ws_right_fields, [{"right_field": f} for f in right_only_fields], headers=["right_field"])

    ws_left_table = wb.create_sheet(_sheet_name(f"left_{left_table}"))
    _write_table(ws_left_table, left_data_norm, headers=left_select_fields)

    ws_right_table = wb.create_sheet(_sheet_name(f"right_{right_table}"))
    _write_table(ws_right_table, right_data_norm, headers=right_select_fields)

    wb.save(out_path)

    trace = {
        "left_sql": left_sql if god_mode else "hidden",
        "right_sql": right_sql if god_mode else "hidden",
        "left_seconds": l_sec,
        "right_seconds": r_sec,
        "left_rows": len(left_data),
        "right_rows": len(right_data),
        "sample_employee_ids": bool(sample_employee_ids and not employee_id and lk_emp and rk_emp),
        "sample_size_requested": int(sample_size),
    }

    return {
        "output_file": str(out_path),
        "counts": {
            "only_in_left": len(only_left_rows),
            "only_in_right": len(only_right_rows),
            "field_differences": len(field_differences_rows),
        },
        "unmapped_fields": {
            "left_only_not_in_field_map": left_only_fields,
            "right_only_not_in_field_map": right_only_fields,
        },
        "trace": trace,
    }


def employee_trace(
    conn: dict,
    table: str,
    employee_id: str,
    fields: list[str],
    employee_field_name: str,
) -> dict[str, Any]:
    cols = ", ".join(_quote_ident(c) for c in fields) if fields else "*"
    sql = f"SELECT {cols} FROM {_quote_table(table)} WHERE {_quote_ident(employee_field_name)} = %s LIMIT 200"
    out_cols, rows, sec = run_query(conn, sql, (employee_id,))
    return {
        "table": table,
        "sql": sql,
        "seconds": sec,
        "columns": out_cols,
        "rows": _rows_to_dicts(out_cols, rows),
    }
