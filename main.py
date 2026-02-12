from __future__ import annotations

import csv
import io
import re
from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook

from app.compare_service import compare_tables, employee_trace
from app.db import get_table_columns
from app.jobs_store import get_job, list_jobs, save_job
from app.mapping_store import load_mapping, field_map_for_table
from app.settings import settings

app = FastAPI(title="Vertica Workday Compare")
templates = Jinja2Templates(directory="templates")


def _conn_left() -> dict:
    return settings.left.model_dump()


def _conn_right() -> dict:
    return settings.right.model_dump()


def _render_home(request: Request, compare_result=None, trace_frames=None, employee_id=""):
    table_maps, field_maps = load_mapping(settings.mapping_file)
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "table_maps": table_maps,
            "field_maps": field_maps,
            "god_mode_default": settings.god_mode_default,
            "jobs": list_jobs(),
            "compare_result": compare_result,
            "trace_frames": trace_frames or [],
            "employee_id": employee_id,
        },
    )


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return _render_home(request)


@app.get("/introspect")
def introspect(left_table: str, right_table: str):
    left_cols = get_table_columns(_conn_left(), left_table)
    right_cols = get_table_columns(_conn_right(), right_table)

    def norm(n: str) -> str:
        return re.sub(r"[^a-z0-9]", "", n.lower())

    right_by_norm = {norm(c): c for c in right_cols}
    suggestions = []
    for lc in left_cols:
        rc = right_by_norm.get(norm(lc))
        if rc:
            suggestions.append({"left_field": lc, "right_field": rc, "confidence": "high"})

    return JSONResponse(
        {
            "left_table": left_table,
            "right_table": right_table,
            "left_columns": left_cols,
            "right_columns": right_cols,
            "suggested_field_mappings": suggestions,
        }
    )


@app.post("/compare", response_class=HTMLResponse)
def compare_ui(
    request: Request,
    left_table: str = Form(...),
    right_table: str = Form(...),
    limit: int = Form(500),
    god_mode: bool = Form(False),
    trim_strings: bool = Form(True),
    nullish_equal: bool = Form(True),
    number_precision: int = Form(6),
):
    _, field_maps = load_mapping(settings.mapping_file)
    result = compare_tables(
        _conn_left(),
        _conn_right(),
        left_table,
        right_table,
        field_maps,
        limit,
        god_mode,
        trim_strings,
        nullish_equal,
        number_precision,
    )
    return _render_home(request, compare_result={"left_table": left_table, "right_table": right_table, "result": result})


@app.post("/compare/export")
def compare_export(
    left_table: str = Form(...),
    right_table: str = Form(...),
    limit: int = Form(500),
    fmt: str = Form("xlsx"),
):
    _, field_maps = load_mapping(settings.mapping_file)
    result = compare_tables(_conn_left(), _conn_right(), left_table, right_table, field_maps, limit, True)

    if fmt == "csv":
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(["section", "payload"])
        for row in result["left_only"]:
            w.writerow(["left_only", row])
        for row in result["right_only"]:
            w.writerow(["right_only", row])
        for row in result["mismatched"]:
            w.writerow(["mismatched", row])
        b = io.BytesIO(out.getvalue().encode("utf-8"))
        return StreamingResponse(b, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=compare_export.csv"})

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "left_only"
    ws1.append(["row_json"])
    for r in result["left_only"]:
        ws1.append([str(r)])
    ws2 = wb.create_sheet("right_only")
    ws2.append(["row_json"])
    for r in result["right_only"]:
        ws2.append([str(r)])
    ws3 = wb.create_sheet("mismatched")
    ws3.append(["row_json"])
    for r in result["mismatched"]:
        ws3.append([str(r)])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=compare_export.xlsx"},
    )


@app.post("/trace", response_class=HTMLResponse)
async def trace_ui(request: Request):
    form = await request.form()
    employee_id = str(form.get("employee_id", "")).strip()
    god_mode = str(form.get("god_mode", "")).lower() in {"on", "true", "1", "yes"}

    table_maps, field_maps = load_mapping(settings.mapping_file)
    selected_tables = form.getlist("left_table")

    frames = []
    for lt in selected_tables:
        tm = next((t for t in table_maps if t.left_table == lt), None)
        if not tm:
            continue
        fmaps = field_map_for_table(field_maps, tm.left_table, tm.right_table)
        l_fields_all = [f.left_field for f in fmaps]
        r_fields_all = [f.right_field for f in fmaps]

        selected_left_fields = form.getlist(f"fields_left::{tm.left_table}")
        selected_right_fields = form.getlist(f"fields_right::{tm.right_table}")

        l_fields = selected_left_fields or l_fields_all
        r_fields = selected_right_fields or r_fields_all

        left_frame = employee_trace(_conn_left(), tm.left_table, employee_id, l_fields, "employee_id")
        right_frame = employee_trace(_conn_right(), tm.right_table, employee_id, r_fields, "employee_id")

        if not god_mode:
            left_frame["sql"] = "hidden"
            right_frame["sql"] = "hidden"

        frames.append({"left_table": tm.left_table, "right_table": tm.right_table, "left": left_frame, "right": right_frame})

    return _render_home(request, trace_frames=frames, employee_id=employee_id)


@app.post("/jobs/save")
def save_job_ui(
    job_type: str = Form(...),
    name: str = Form(...),
    left_table: str = Form(""),
    right_table: str = Form(""),
    limit: int = Form(500),
):
    payload = {"left_table": left_table, "right_table": right_table, "limit": limit}
    save_job(job_type, name, payload)
    return RedirectResponse("/", status_code=303)


@app.post("/jobs/{job_id}/run")
def run_job(job_id: int):
    j = get_job(job_id)
    if not j:
        return JSONResponse({"error": "job not found"}, status_code=404)
    return JSONResponse({"job": j, "message": "Use payload values in /compare or /trace forms (replay wiring ready)."})
