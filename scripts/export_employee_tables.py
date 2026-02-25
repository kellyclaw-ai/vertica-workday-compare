"""Export raw per-table employee_id extracts to an Excel workbook.

- Interactive CLI table selection.
- Runs SELECT * for each selected table.
- Interactive data scope options:
  - single-column filter per table (recommended)
  - random/semi-random 100-row sample
  - full table
- Sorts by effective date when present (effective_dt or effective_date).
- Highlights cells whose values changed vs the previous record (by effective date).

Usage:
  python3 scripts/export_employee_tables.py
  python3 scripts/export_employee_tables.py --out out.xlsx

Notes:
- This script uses the same Vertica connection settings as the FastAPI app (app/settings.py).
- For table listing it prefers schema enumeration; if schema isn't provided it tries to infer it
  from the mapping workbook.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.db import get_schema_tables, get_table_columns, run_query
from app.mapping_store import load_mapping
from app.settings import settings


MISMATCH_FILL = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")


def _sheet_name(name: str) -> str:
    safe = name.replace("/", "_").replace("\\", "_").replace(".", "_")
    return safe[:31] if len(safe) > 31 else safe


def _unique_sheet_name(wb: Workbook, base_name: str) -> str:
    """Return an Excel-safe unique sheet name (<=31 chars)."""
    base = _sheet_name(base_name)
    if base not in wb.sheetnames:
        return base

    i = 2
    while True:
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        if candidate not in wb.sheetnames:
            return candidate
        i += 1


def _quote_ident(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def _quote_table(table_name: str) -> str:
    return ".".join(_quote_ident(part) for part in table_name.split("."))


def _norm(v: Any):
    # Treat blank/whitespace strings as NULL
    if v is None:
        return None
    if isinstance(v, str):
        sv = v.strip()
        return None if sv == "" else sv
    return v


def _distinct(a: Any, b: Any) -> bool:
    na = _norm(a)
    nb = _norm(b)
    if na is None and nb is None:
        return False
    return na != nb


def _find_effective_col(columns: list[str]) -> str | None:
    for c in columns:
        cl = c.lower()
        if cl in {"effective_dt", "effective_date"}:
            return c
    return None


def _sortable_dt(v: Any):
    if v is None:
        return datetime.min
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return datetime.min
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
            "%m/%d/%Y",
        ):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass
        try:
            return datetime.fromisoformat(s.replace("Z", "+00:00"))
        except Exception:
            return datetime.min
    return datetime.min


def _excel_safe(v: Any):
    # openpyxl cannot write some native Python objects (e.g., uuid.UUID)
    if isinstance(v, UUID):
        return str(v)
    if isinstance(v, (datetime, date, int, float, bool, str)) or v is None:
        return v
    # Fallback for Decimal, bytes, custom objects, etc.
    return str(v)


def _write_table(ws, columns: list[str], rows: list[dict[str, Any]], *, changed_cols_by_row: list[set[str]] | None = None):
    ws.append(columns)
    for i, r in enumerate(rows):
        ws.append([_excel_safe(r.get(c)) for c in columns])
        if changed_cols_by_row is None:
            continue
        changed = changed_cols_by_row[i] if i < len(changed_cols_by_row) else set()
        if not changed:
            continue
        # Row in worksheet is offset by 1 for header
        row_idx = i + 2
        for j, c in enumerate(columns):
            if c in changed:
                ws.cell(row=row_idx, column=j + 1).fill = MISMATCH_FILL


def _parse_selection(selection: str, n: int) -> list[int]:
    """Parse '1,2,5-8,all' into 0-based indices."""
    s = selection.strip().lower()
    if not s:
        return []
    if s in {"a", "all", "*"}:
        return list(range(n))

    out: set[int] = set()
    for part in s.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            try:
                start = int(a)
                end = int(b)
            except ValueError:
                continue
            for k in range(min(start, end), max(start, end) + 1):
                if 1 <= k <= n:
                    out.add(k - 1)
        else:
            try:
                k = int(part)
            except ValueError:
                continue
            if 1 <= k <= n:
                out.add(k - 1)
    return sorted(out)


@dataclass
class TableRef:
    side: str  # left|right
    table: str  # fully-qualified schema.table


@dataclass
class TableFilter:
    column: str
    value: str


def _infer_schema_from_mapping(which: str) -> str | None:
    # Infer schema from the mapping file (first mapped table that contains a schema prefix)
    table_maps, _field_maps, _value_maps = load_mapping(settings.mapping_file)
    cand = None
    if which == "left":
        cand = next((tm.left_table for tm in table_maps if tm.left_table and "." in tm.left_table), None)
    else:
        cand = next((tm.right_table for tm in table_maps if tm.right_table and "." in tm.right_table), None)
    if cand and "." in cand:
        return cand.split(".", 1)[0]
    return None


def _effective_like(col: str) -> bool:
    c = col.lower()
    return c in {"effective_dt", "effective_date"} or "effective" in c


def _rank_filter_columns(columns: list[str], key_candidates: list[str]) -> list[str]:
    """Prioritize mapped key fields, but de-prioritize effective-date keys."""
    existing = set(columns)
    ranked: list[str] = []

    # 1) mapped keys that exist on the table, non-effective first
    mapped = [c for c in key_candidates if c in existing]
    mapped_non_eff = [c for c in mapped if not _effective_like(c)]
    mapped_eff = [c for c in mapped if _effective_like(c)]
    ranked.extend(mapped_non_eff)
    ranked.extend(mapped_eff)

    # 2) common id-style columns from schema
    id_like = [c for c in columns if c.lower().endswith("_id") and c not in ranked]
    id_like_non_eff = [c for c in id_like if not _effective_like(c)]
    id_like_eff = [c for c in id_like if _effective_like(c)]
    ranked.extend(id_like_non_eff)
    ranked.extend(id_like_eff)

    # 3) everything else, non-effective then effective
    remaining = [c for c in columns if c not in ranked]
    remaining_non_eff = [c for c in remaining if not _effective_like(c)]
    remaining_eff = [c for c in remaining if _effective_like(c)]
    ranked.extend(remaining_non_eff)
    ranked.extend(remaining_eff)
    return ranked


def _list_tables(conn: dict, schema: str) -> list[str]:
    tables = get_schema_tables(conn, schema)
    # app.db.get_schema_tables already returns fully-qualified schema.table names.
    # Keep a defensive normalization in case implementation changes.
    out: list[str] = []
    for t in tables:
        out.append(t if "." in t else f"{schema}.{t}")
    return out


def _trace_explorer_table_refs() -> list[TableRef]:
    table_maps, _field_maps, _value_maps = load_mapping(settings.mapping_file)

    left_tables = sorted({tm.left_table for tm in table_maps if tm.left_table})

    right_schema = _infer_schema_from_mapping("right")
    right_tables: list[str] = []
    if right_schema:
        try:
            right_tables = _list_tables(settings.right.model_dump(), right_schema)
        except Exception:
            right_tables = []
    if not right_tables:
        right_tables = sorted({tm.right_table for tm in table_maps if tm.right_table})

    return ([TableRef("left", t) for t in left_tables] + [TableRef("right", t) for t in right_tables])


def _mapping_key_candidates_for_ref(ref: TableRef) -> list[str]:
    _table_maps, field_maps, _value_maps = load_mapping(settings.mapping_file)
    keys: list[str] = []
    if ref.side == "left":
        for fm in field_maps:
            if fm.left_table == ref.table and fm.is_key and fm.left_field:
                keys.append(fm.left_field)
    else:
        for fm in field_maps:
            if fm.right_table == ref.table and fm.is_key and fm.right_field:
                keys.append(fm.right_field)

    # preserve order while removing duplicates
    out: list[str] = []
    seen: set[str] = set()
    for k in keys:
        if k not in seen:
            out.append(k)
            seen.add(k)
    return out


def _prompt_table_filter(ref: TableRef, conn: dict) -> TableFilter | None:
    cols = get_table_columns(conn, ref.table)
    if not cols:
        print(f"\n[{ref.side}] {ref.table}: no columns found; skipping filter.")
        return None

    key_candidates = _mapping_key_candidates_for_ref(ref)
    ranked = _rank_filter_columns(cols, key_candidates)

    print(f"\nFilter for [{ref.side}] {ref.table}")
    print("Choose one column to filter this table (or press Enter to skip):")
    for i, c in enumerate(ranked, start=1):
        tag = ""
        if c in key_candidates:
            tag = " [key]"
        if _effective_like(c):
            tag += " (effective-date-like)"
        print(f"  {i:>3}) {c}{tag}")

    while True:
        raw = input("Column number (Enter=skip): ").strip()
        if raw == "":
            return None
        try:
            idx = int(raw)
        except ValueError:
            print("Invalid selection; try again.")
            continue
        if not (1 <= idx <= len(ranked)):
            print("Out of range; try again.")
            continue
        col = ranked[idx - 1]
        break

    value = ""
    while value == "":
        value = input(f"Value for {col}: ").strip()
        if value == "":
            print("Value is required (or skip by re-selecting column as Enter).")

    return TableFilter(column=col, value=value)


def export_table_ref(
    wb: Workbook,
    ref: TableRef,
    *,
    scope: str,
    table_filter: TableFilter | None = None,
    sample_size: int = 100,
):
    conn = settings.left.model_dump() if ref.side == "left" else settings.right.model_dump()
    table = ref.table

    cols = get_table_columns(conn, table)
    eff_col = _find_effective_col(cols)
    eff_order_sql = f" ORDER BY {_quote_ident(eff_col)} ASC" if eff_col else ""

    if scope == "filtered":
        if not table_filter:
            raise ValueError("table_filter is required for filtered scope")
        sql = (
            f"SELECT * FROM {_quote_table(table)} "
            f"WHERE {_quote_ident(table_filter.column)} = %s{eff_order_sql}"
        )
        params: tuple[Any, ...] | None = (table_filter.value,)
    elif scope == "sample100":

        # Semi-random sample using ORDER BY RANDOM(). If effective date exists, re-sort sampled set by effective date.
        if eff_col:
            sql = (
                f"SELECT * FROM ("
                f"SELECT * FROM {_quote_table(table)} ORDER BY RANDOM() LIMIT {int(sample_size)}"
                f") s ORDER BY {_quote_ident(eff_col)} ASC"
            )
        else:
            sql = f"SELECT * FROM {_quote_table(table)} ORDER BY RANDOM() LIMIT {int(sample_size)}"
        params = None
    elif scope == "full":
        sql = f"SELECT * FROM {_quote_table(table)}{eff_order_sql}"
        params = None
    else:
        raise ValueError(f"Unsupported scope: {scope}")

    out_cols, out_rows, _sec = run_query(conn, sql, params)
    rows = [dict(zip(out_cols, r)) for r in out_rows]

    # Defensive sort in Python too
    if eff_col and eff_col in out_cols:
        rows = sorted(rows, key=lambda r: _sortable_dt(r.get(eff_col)))

    # Build per-row changed columns (relative to previous row)
    changed_cols_by_row: list[set[str]] = []
    ignore = {"employee_id", "effective_dt", "effective_date"}
    prev: dict[str, Any] | None = None
    for r in rows:
        changed: set[str] = set()
        if prev is not None:
            for c in out_cols:
                if c.lower() in ignore:
                    continue
                if _distinct(r.get(c), prev.get(c)):
                    changed.add(c)
        changed_cols_by_row.append(changed)
        prev = r

    # Sheet name: table name only (no side/schema), truncated to Excel limits.
    # If duplicate table names exist across sides/schemas, auto-deduplicate with numeric suffix.
    table_only = table.split(".")[-1]
    ws = wb.create_sheet(_unique_sheet_name(wb, table_only))
    _write_table(ws, out_cols, rows, changed_cols_by_row=changed_cols_by_row)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=None, help="Output xlsx path")
    args = ap.parse_args()

    print("\nData scope:")
    print("  1) Single-column filter per table (recommended)")
    print("  2) Random/semi-random 100 rows")
    print("  3) Full table")

    scope_choice = ""
    while scope_choice not in {"1", "2", "3"}:
        scope_choice = input("Choose scope [1/2/3]: ").strip()

    scope = {"1": "filtered", "2": "sample100", "3": "full"}[scope_choice]

    refs = _trace_explorer_table_refs()
    if not refs:
        raise SystemExit("No tables found (mapping file empty?)")

    print(f"Found {len(refs)} tables (same list as trace explorer).")
    for i, r in enumerate(refs, start=1):
        side_tag = "L" if r.side == "left" else "R"
        print(f"{i:>4}: [{side_tag}] {r.table}")

    selected: list[TableRef] = []
    while True:
        raw = input("\nSelect tables (e.g. 1,2,5-7 or 'all'). Press Enter when done: ").strip()
        if raw == "":
            break
        idxs = _parse_selection(raw, len(refs))
        if not idxs:
            print("No valid selections parsed; try again.")
            continue
        for ix in idxs:
            r = refs[ix]
            if all((r.side != s.side or r.table != s.table) for s in selected):
                selected.append(r)
        print(f"Selected so far ({len(selected)}): {', '.join([f'{s.side}:{s.table}' for s in selected[:6]])}{' ...' if len(selected) > 6 else ''}")

    if not selected:
        raise SystemExit("No tables selected.")

    table_filters: dict[tuple[str, str], TableFilter | None] = {}
    if scope == "filtered":
        print("\nYou will choose one filter column/value per selected table.")
        for ref in selected:
            conn = settings.left.model_dump() if ref.side == "left" else settings.right.model_dump()
            table_filters[(ref.side, ref.table)] = _prompt_table_filter(ref, conn)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    scope_label = "filtered" if scope == "filtered" else ("sample100" if scope == "sample100" else "full")
    out_path = Path(args.out) if args.out else Path(f"output/table_extract_{scope_label}_{ts}.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    exported = 0
    for ref in selected:
        if scope == "filtered" and table_filters.get((ref.side, ref.table)) is None:
            print(f"Skipping [{ref.side}] {ref.table} (no filter selected).")
            continue
        export_table_ref(
            wb,
            ref,
            scope=scope,
            table_filter=table_filters.get((ref.side, ref.table)),
            sample_size=100,
        )
        exported += 1

    if exported == 0:
        raise SystemExit("No sheets exported. Nothing to write.")

    wb.save(out_path)
    print(f"\nWrote workbook: {out_path}")


if __name__ == "__main__":
    main()
