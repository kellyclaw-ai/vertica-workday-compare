"""Compare two Excel datasets (same schema) by WorkdayID and export summary workbook.

What it does:
- Prompts for left/right Excel file paths in CLI.
- Reads the first sheet from each workbook.
- Uses WorkdayID as the primary key.
- Computes row counts:
  - left only
  - right only
  - both
- For rows present on both sides, compares each shared field and counts differences.
- Writes output workbook to ./output with tabs:
  - Left only row count
  - Right only row count
  - Both row count
  - Field difference summary
  - Compared files

Usage:
  python3 scripts/compare_excel_files.py
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

KEY_FIELD = "WorkdayID"


def _sheet_name(name: str) -> str:
    safe = name.replace("/", "_").replace("\\", "_").replace(".", "_")
    return safe[:31] if len(safe) > 31 else safe


def _norm(v: Any):
    # Treat blank/whitespace as null-equivalent for comparison
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return None if s == "" else s
    return v


def _read_first_sheet(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return [], []

    headers = ["" if h is None else str(h).strip() for h in header_row]
    data: list[dict[str, Any]] = []

    for r in rows:
        rec = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
        # Skip completely blank rows
        if any(v is not None and str(v).strip() != "" for v in rec.values()):
            data.append(rec)

    return headers, data


def _to_index(rows: list[dict[str, Any]], key_field: str) -> dict[Any, dict[str, Any]]:
    out: dict[Any, dict[str, Any]] = {}
    for r in rows:
        k = _norm(r.get(key_field))
        if k is None:
            continue
        # If duplicate keys exist, keep the last seen record.
        out[k] = r
    return out


def _write_simple_count_tab(wb: Workbook, title: str, count: int):
    ws = wb.create_sheet(_sheet_name(title))
    ws.append(["metric", "value"])
    ws.append([title, count])


def main():
    left_raw = input("Enter LEFT Excel file path: ").strip()
    right_raw = input("Enter RIGHT Excel file path: ").strip()

    if not left_raw or not right_raw:
        raise SystemExit("Both file paths are required.")

    left_path = Path(left_raw).expanduser().resolve()
    right_path = Path(right_raw).expanduser().resolve()

    if not left_path.exists():
        raise SystemExit(f"Left file not found: {left_path}")
    if not right_path.exists():
        raise SystemExit(f"Right file not found: {right_path}")

    left_headers, left_rows = _read_first_sheet(left_path)
    right_headers, right_rows = _read_first_sheet(right_path)

    if KEY_FIELD not in left_headers:
        raise SystemExit(f"Missing required key field '{KEY_FIELD}' in LEFT file header.")
    if KEY_FIELD not in right_headers:
        raise SystemExit(f"Missing required key field '{KEY_FIELD}' in RIGHT file header.")

    # Use shared fields only; skip key in field-level diff summary.
    shared_fields = [h for h in left_headers if h in set(right_headers)]
    compare_fields = [h for h in shared_fields if h != KEY_FIELD]

    left_ix = _to_index(left_rows, KEY_FIELD)
    right_ix = _to_index(right_rows, KEY_FIELD)

    left_keys = set(left_ix.keys())
    right_keys = set(right_ix.keys())

    left_only_keys = left_keys - right_keys
    right_only_keys = right_keys - left_keys
    both_keys = left_keys & right_keys

    # Field difference counts across rows on both sides
    diff_counts: dict[str, int] = {f: 0 for f in compare_fields}
    for k in both_keys:
        lrow = left_ix[k]
        rrow = right_ix[k]
        for f in compare_fields:
            if _norm(lrow.get(f)) != _norm(rrow.get(f)):
                diff_counts[f] += 1

    out_dir = Path("output")
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"excel_compare_{left_path.stem}_vs_{right_path.stem}_{ts}.xlsx"

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    _write_simple_count_tab(wb_out, "Left only row count", len(left_only_keys))
    _write_simple_count_tab(wb_out, "Right only row count", len(right_only_keys))
    _write_simple_count_tab(wb_out, "Both row count", len(both_keys))

    ws_summary = wb_out.create_sheet(_sheet_name("Field difference summary"))
    ws_summary.append(["field", "difference_count"])
    for f, c in sorted(diff_counts.items(), key=lambda x: (-x[1], x[0].lower())):
        ws_summary.append([f, c])

    ws_files = wb_out.create_sheet(_sheet_name("Compared files"))
    ws_files.append(["side", "file_name", "full_path"])
    ws_files.append(["left", left_path.name, str(left_path)])
    ws_files.append(["right", right_path.name, str(right_path)])

    wb_out.save(out_path)
    print(f"Wrote comparison workbook: {out_path}")


if __name__ == "__main__":
    main()
